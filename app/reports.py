from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable, List, Optional, Tuple, Dict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _autosize_columns(ws, max_col: int):
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def build_payments_xlsx(rows: List[dict], title: str) -> bytes:
    """
    rows: [{"id":..,"created_at":..,"tg_id":..,"pay_code":..,"provider":..,"amount":..,"status":..,"plan_days":..,"ext_id":..}, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "payments"

    ws.append([title])
    ws.append(["Generated (UTC)", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])

    headers = [
        "payment_id",
        "created_at_utc",
        "tg_id",
        "pay_code",
        "provider",
        "amount_uzs",
        "status",
        "plan_days",
        "ext_id",
    ]
    ws.append(headers)

    for r in rows:
        created = r.get("created_at")
        created_str = created.strftime("%Y-%m-%d %H:%M:%S") if created else ""
        ws.append([
            r.get("id"),
            created_str,
            r.get("tg_id"),
            r.get("pay_code"),
            r.get("provider"),
            r.get("amount"),
            r.get("status"),
            r.get("plan_days"),
            r.get("ext_id"),
        ])

    _autosize_columns(ws, len(headers))

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def payments_stats(rows: List[dict]) -> Dict[str, dict]:
    """
    returns:
      {"payme": {"count":..,"sum":..}, "click":..., "admin":..., "all": {...}}
    """
    by: Dict[str, dict] = {}
    total_count = 0
    total_sum = 0

    for r in rows:
        provider = (r.get("provider") or "unknown").lower()
        amount = int(r.get("amount") or 0)

        by.setdefault(provider, {"count": 0, "sum": 0})
        by[provider]["count"] += 1
        by[provider]["sum"] += amount

        total_count += 1
        total_sum += amount

    by["all"] = {"count": total_count, "sum": total_sum}
    return by
